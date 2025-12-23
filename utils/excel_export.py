import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_excel(df: pd.DataFrame) -> bytes:
    """
    Genera un archivo Excel en memoria con formato profesional.
    - Encabezados con color.
    - Autofiltro.
    - Paneles inmovilizados.
    - Ajuste de columnas.
    - Resaltado de comunas de Detracción.
    """
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Cobranzas"

    # Estilos
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid") # Azul Antay
    header_font = Font(color="FFFFFF", bold=True)
    
    detra_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Amarillo suave
    estado_dt_fill_pendiente = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Rojo suave
    estado_dt_fill_pagado = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid") # Verde suave

    # Escribir encabezados en MAYÚSCULAS
    headers = [str(h).upper() for h in df.columns]
    ws.append(headers)
    
    # Formato encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Inmovilizar fila 1
    ws.freeze_panes = "A2"
    
    # Autofiltro
    ws.auto_filter.ref = ws.dimensions

    # Identificar índice de columna MONEDA para formato condicional
    try:
        moneda_idx = headers.index("MONEDA")
    except ValueError:
        moneda_idx = -1

    def _clean_currency(val):
        """Convierte string de moneda a float."""
        if isinstance(val, (int, float)):
            return float(val)
        if not val or not isinstance(val, str):
            return 0.0
        # Remover símbolos comunes
        clean_val = val.replace('S/', '').replace('$', '').replace(',', '').replace(' ', '')
        try:
            return float(clean_val)
        except ValueError:
            return 0.0

    # Escribir datos
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        # Obtener moneda de la fila actual (si existe columna MONEDA)
        current_currency = "S/"
        if moneda_idx != -1 and len(row) > moneda_idx:
            val_mon = str(row[moneda_idx]).strip().upper()
            if "$" in val_mon or "DOL" in val_mon:
                current_currency = "$"

        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Columnas clave (Upper para coincidir con headers)
            col_name = headers[c_idx - 1]
            
            # Formato de Fechas
            if "FECH" in col_name and value:
                cell.number_format = 'DD/MM/YYYY'
                
            # Formato de Moneda y Números
            # Incluir nuevas columnas: SALDO REAL, IMPORTE REFERENCIAL (S/), TIPO CAMBIO
            # [FIX RC-BUG-003] AMORTIZACIONES es TEXTO, sacar de limpieza numérica
            monetary_cols = [
                "MONT EMIT", "CALCULADO (S/)", "DETRACCIÓN", "SALDO", 
                "SALDO REAL", "IMPORTE REFERENCIAL (S/)", "TIPO CAMBIO"
            ]
            
            if col_name in monetary_cols:
                # 1. Convertir a número (float)
                num_val = _clean_currency(value)
                cell.value = num_val
                
                # 2. Aplicar formato según Moneda
                if col_name == "TIPO CAMBIO":
                     cell.number_format = '0.000'
                elif col_name == "DETRACCIÓN":
                     # [FIX RC-BUG-002] Detracción SIEMPRE es en Soles
                     cell.number_format = '"S/" #,##0.00'
                else:
                    if current_currency == "$":
                        cell.number_format = '"$" #,##0.00'
                    else:
                        cell.number_format = '"S/" #,##0.00'
            
            # Ajuste de Texto para columnas con saltos de línea
            if col_name in ["ESTADO DETRACCION", "AMORTIZACIONES"]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Resaltar Columnas Detracción
            if col_name == "DETRACCIÓN":
                cell.fill = detra_fill
                
            if col_name == "ESTADO DETRACCION":
                if value == "Pendiente":
                     cell.fill = estado_dt_fill_pendiente
                elif value != "-" and value != "No Aplica":
                     cell.fill = estado_dt_fill_pagado

    # Ajuste de ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells[:20]) # Sample primeros 20
        length = min(length, 50) # Max ancho 50
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(output)
    return output.getvalue()
