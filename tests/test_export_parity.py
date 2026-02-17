from io import BytesIO

from openpyxl import load_workbook

from tests.fixtures.synthetic_data import fixture_fresh_load
from utils.excel_export import (
    EXPORT_COLUMNS_BASELINE,
    build_export_dataframe,
    generate_excel,
    get_export_columns,
)

# Baseline funcional heredado del flujo pre-migracion (lista legacy de export).
LEGACY_EXPORT_COLUMNS_BASELINE = [
    'COD CLIENTE',
    'EMPRESA',
    'Enviar Email',
    'ESTADO_EMAIL',
    'FECHA_ULTIMO_ENVIO',
    'NOTA',
    'CORREO',
    'TELÉFONO',
    'TIPO PEDIDO',
    'COMPROBANTE',
    'FECH EMIS',
    'FECH VENC',
    'DÍAS MORA',
    'ESTADO DEUDA',
    'MONEDA',
    'TIPO CAMBIO',
    'MONT EMIT',
    'DETRACCIÓN',
    'ESTADO DETRACCION',
    'AMORTIZACIONES',
    'SALDO',
    'SALDO REAL',
    'MATCH_KEY',
]


def _df_with_full_export_schema():
    df = fixture_fresh_load().copy()
    # Fixture legacy no siempre trae todas las columnas del contrato de export.
    for col in EXPORT_COLUMNS_BASELINE:
        if col not in df.columns:
            df[col] = ""
    if "Enviar Email" in df.columns:
        df["Enviar Email"] = "SI"
    return df


def test_export_contract_matches_legacy_baseline():
    assert EXPORT_COLUMNS_BASELINE == LEGACY_EXPORT_COLUMNS_BASELINE


def test_export_columns_match_baseline_order():
    df = _df_with_full_export_schema()
    export_cols = get_export_columns(df)
    assert export_cols == EXPORT_COLUMNS_BASELINE


def test_build_export_dataframe_keeps_order_and_index():
    df = _df_with_full_export_schema()
    df_export = build_export_dataframe(df)

    assert list(df_export.columns) == EXPORT_COLUMNS_BASELINE
    assert list(df_export.index) == list(range(1, len(df_export) + 1))


def test_generate_excel_keeps_header_order_uppercase():
    df = _df_with_full_export_schema().head(1)
    df_export = build_export_dataframe(df)
    excel_bytes = generate_excel(df_export)

    wb = load_workbook(filename=BytesIO(excel_bytes))
    ws = wb.active
    headers = [ws.cell(row=1, column=i + 1).value for i in range(len(EXPORT_COLUMNS_BASELINE))]
    assert headers == [col.upper() for col in EXPORT_COLUMNS_BASELINE]


def test_generate_excel_currency_format_by_moneda_and_detraction():
    df = _df_with_full_export_schema().head(2).copy()
    df.loc[df.index[0], "MONEDA"] = "SOLES"
    df.loc[df.index[1], "MONEDA"] = "DOLARES"
    df.loc[:, "MONT EMIT"] = [1000.0, 2000.0]
    df.loc[:, "SALDO"] = [900.0, 1800.0]
    df.loc[:, "SALDO REAL"] = [800.0, 1700.0]
    df.loc[:, "DETRACCIÓN"] = [120.0, 240.0]

    df_export = build_export_dataframe(df)
    excel_bytes = generate_excel(df_export)

    wb = load_workbook(filename=BytesIO(excel_bytes))
    ws = wb.active

    saldo_real_col = EXPORT_COLUMNS_BASELINE.index("SALDO REAL") + 1
    detraccion_col = EXPORT_COLUMNS_BASELINE.index("DETRACCIÓN") + 1

    # Row 2 = moneda SOLES, row 3 = moneda DOLARES.
    assert ws.cell(row=2, column=saldo_real_col).number_format == '"S/" #,##0.00'
    assert ws.cell(row=3, column=saldo_real_col).number_format == '"$" #,##0.00'
    # Detraccion siempre en Soles.
    assert ws.cell(row=2, column=detraccion_col).number_format == '"S/" #,##0.00'
    assert ws.cell(row=3, column=detraccion_col).number_format == '"S/" #,##0.00'


def test_export_dataframe_respects_filtered_subset():
    df = _df_with_full_export_schema()
    filtered = df[df["EMPRESA"].isin(["Empresa 1", "Empresa 3", "Empresa 5"])].copy()

    df_export = build_export_dataframe(filtered)

    assert len(df_export) == len(filtered)
    assert set(df_export["EMPRESA"].unique()) == set(filtered["EMPRESA"].unique())
